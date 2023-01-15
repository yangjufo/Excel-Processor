import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# file class


class BaseFile():
    def __init__(self, path) -> None:
        self.path = path
        self._headers = list()
        self._column_values_map = dict()

    def get_sheet_names(self):
        return [True, []]

    def get_headers(self, sheet_name, read_headers):
        if len(self._headers) == 0:
            # read headers from child class
            headers_frame = read_headers(sheet_name)
            index = 0
            for c in headers_frame.values.tolist()[0]:
                self._headers.append(str(index) + " " + str(c))
                index += 1
        return [True, self._headers]

    def get_column_values(self, sheet_name, column_index, exclude_header, read_column_values):
        if column_index not in self._column_values_map:
            # read column values from child class
            column_values_frame = read_column_values(
                sheet_name, column_index)
            cols = set()
            for c in column_values_frame.values.tolist()[1:]:
                if isinstance(c[0], datetime):
                    cols.add(c[0].strftime('%Y-%m-%d'))
                else:
                    cols.add(str(c[0]))
            self._column_values_map[column_index] = sorted(cols)
        cols = self._column_values_map[column_index]
        if not exclude_header:
            cols.append(self._headers[column_index].split(' ', 1)[1])
        return [True, sorted(cols)]


# CSV file
class CsvFile(BaseFile):
    def __init__(self, path) -> None:
        super().__init__(path)

    def get_headers(self, sheet_name):
        return super().get_headers(sheet_name, self.__read_headers)

    def __read_headers(self, _):
        return pd.read_csv(self.path, nrows=1, header=None)

    def get_column_values(self, sheet_name, column_index, exclude_header):
        return super().get_column_values(sheet_name, column_index, exclude_header, self.__read_column_values)

    def __read_column_values(self, _, column_index):
        return pd.read_csv(self.path, header=None, usecols=[column_index])

    def get_row_values(self, _, nrows):
        return pd.read_csv(self.path, nrows=nrows)

    def get_all_row_values(self, _):
        return pd.read_csv(self.path)


# XLS, XLSX file
class XlsFile(BaseFile):
    def __init__(self, path) -> None:
        super().__init__(path)
        self._sheet_names = list()

    def get_sheet_names(self):
        if len(self._sheet_names) == 0:
            self._sheet_names = load_workbook(
                self.path, read_only=True).sheetnames
        return [True, self._sheet_names]

    def get_headers(self, sheet_name):
        return super().get_headers(sheet_name, self.__read_headers)

    def __read_headers(self, sheet_name):
        return pd.read_excel(
            self.path, sheet_name=sheet_name, nrows=1, header=None)

    def get_column_values(self, sheet_name, column_index, exclude_header):
        return super().get_column_values(sheet_name, column_index, exclude_header, self.__read_column_values)

    def __read_column_values(self, sheet_name, column_index):
        return pd.read_excel(
            self.path, sheet_name=sheet_name, header=None, usecols=[column_index])

    def get_row_values(self, sheet_name, nrows):
        return pd.read_excel(self.path, sheet_name=sheet_name, nrows=nrows)

    def get_all_row_values(self, sheet_name):
        return pd.read_excel(self.path, sheet_name=sheet_name)
